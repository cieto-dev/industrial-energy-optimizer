"""
excel_report.py — Excel report generation (Sprint 3.4).

Purpose
-------
Generate a human-readable Excel report from a Recommendation object,
including a scenario comparison table with multiple technology options.

Key Requirements
----------------
- Scenario comparison table (multiple technologies side by side)
- All numbers must have plain-language explanations
- estimated_total_benefit_inr must show total_benefit_verified status and disclaimer
- Disclaimer must be visible in the Excel file (cell comment or separate row)
- No raw JSON dumps or unexplained jargon
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional, Dict, List
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from models.recommendation import Recommendation
from decision_engine.optimizer.optimization_engine import OptimizationResult


class ExcelReportGenerator:
    """Generate Excel reports from Recommendation objects."""

    def __init__(self, output_path: Path):
        self.output_path = output_path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Recommendation Report"

        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.title_font = Font(bold=True, size=14, color="000000")
        self.section_font = Font(bold=True, size=11, color="000000")
        self.disclaimer_font = Font(italic=True, size=9, color="FF0000")
        self.warning_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        self.center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _format_currency(self, value: float) -> str:
        """Format currency value in INR with plain language."""
        if value >= 1_00_00_000:  # 1 crore
            crores = value / 1_00_00_000
            return f"₹{crores:.2f} crore (₹{value:,.0f})"
        elif value >= 1_00_000:  # 1 lakh
            lakhs = value / 1_00_000
            return f"₹{lakhs:.2f} lakh (₹{value:,.0f})"
        else:
            return f"₹{value:,.0f}"

    def _format_percentage(self, value: float) -> str:
        """Format percentage with plain language."""
        return f"{value:.1f}%"

    def _format_years(self, value: float) -> str:
        """Format years with plain language."""
        if value < 1:
            months = value * 12
            return f"{months:.0f} months ({value:.2f} years)"
        return f"{value:.1f} years"

    def _set_column_width(self, column: str, width: float):
        """Set column width for better readability."""
        self.ws.column_dimensions[column].width = width

    def _merge_and_write(self, row: int, start_col: int, end_col: int, value: str, font=None, fill=None, alignment=None):
        """Merge cells and write value with styling."""
        self.ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        cell = self.ws.cell(row=row, column=start_col, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment

    def generate(self, recommendation: Recommendation, optimization_result: Optional[OptimizationResult] = None) -> Path:
        """
        Generate an Excel report from the Recommendation.

        Parameters
        ----------
        recommendation : Recommendation
            The recommendation object to convert to Excel
        optimization_result : OptimizationResult, optional
            Full optimization result for scenario comparison table

        Returns
        -------
        Path
            Path to the generated Excel file
        """
        row = 1

        # Build report sections
        row = self._build_title_section(recommendation, row)
        row = self._build_summary_section(recommendation, row)
        row = self._build_recommendation_section(recommendation, row)
        row = self._build_economic_section(recommendation, row)
        row = self._build_environmental_section(recommendation, row)
        row = self._build_policy_section(recommendation, row)
        row = self._build_sensitivity_section(recommendation, row)

        # Add scenario comparison table if optimization result is available
        if optimization_result:
            row = self._build_scenario_comparison_table(recommendation, optimization_result, row)

        row = self._build_alternatives_section(recommendation, row)
        row = self._build_disclaimer_section(recommendation, row)

        # Set column widths
        self._set_column_width('A', 25)
        self._set_column_width('B', 30)
        self._set_column_width('C', 25)
        self._set_column_width('D', 20)
        self._set_column_width('E', 20)
        self._set_column_width('F', 20)

        self.wb.save(str(self.output_path))
        return self.output_path

    def _build_title_section(self, recommendation: Recommendation, start_row: int) -> int:
        """Build the title section of the report."""
        row = start_row

        # Main title
        self._merge_and_write(
            row, 1, 6,
            f"Industrial Energy Transition Recommendation - {recommendation.factory_name}",
            self.title_font, alignment=self.center_alignment
        )
        row += 1

        # Subtitle
        subtitle = (
            f"Industry: {recommendation.industry} | State: {recommendation.state} | "
            f"Generated: {recommendation.generated_at.strftime('%Y-%m-%d %H:%M')}"
        )
        self._merge_and_write(
            row, 1, 6, subtitle, Font(size=10), alignment=self.center_alignment
        )
        row += 2

        return row

    def _build_summary_section(self, recommendation: Recommendation, start_row: int) -> int:
        """Build the executive summary section."""
        row = start_row

        # Section header
        self._merge_and_write(
            row, 1, 6, "EXECUTIVE SUMMARY", self.section_font, alignment=self.left_alignment
        )
        row += 1

        # Technology sequence
        tech_sequence = " → ".join(recommendation.recommended_technology_sequence)
        self.ws.cell(row=row, column=1, value="Recommended Pathway:")
        self.ws.cell(row=row, column=2, value=tech_sequence)
        self.ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # MCDA summary
        mcda_summary = (
            f"Overall Score: {recommendation.composite_score:.2f} | "
            f"Cost: {recommendation.objective_scores['cost']:.2f} | "
            f"Emissions: {recommendation.objective_scores['emissions']:.2f} | "
            f"Risk: {recommendation.objective_scores['risk']:.2f}"
        )
        self.ws.cell(row=row, column=1, value="MCDA Scores:")
        self.ws.cell(row=row, column=2, value=mcda_summary)
        self.ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # Cost ranking note
        cost_note = (
            "This recommendation is also the most cost-effective option."
            if recommendation.recommended_is_cheapest
            else "This recommendation prioritizes environmental benefits and risk reduction over lowest cost."
        )
        self.ws.cell(row=row, column=1, value="Cost Ranking:")
        self.ws.cell(row=row, column=2, value=cost_note)
        self.ws.cell(row=row, column=1).font = Font(bold=True)
        row += 2

        return row

    def _build_recommendation_section(self, recommendation: Recommendation, start_row: int) -> int:
        """Build the detailed recommendation reasoning section."""
        row = start_row

        # Section header
        self._merge_and_write(
            row, 1, 6, "WHY THIS PATHWAY WAS RECOMMENDED", self.section_font, alignment=self.left_alignment
        )
        row += 1

        # List each reason
        for i, reason in enumerate(recommendation.explanation.why_selected, 1):
            self.ws.cell(row=row, column=1, value=f"{i}.")
            self.ws.cell(row=row, column=2, value=reason)
            self.ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

        row += 1
        return row

    def _build_economic_section(self, recommendation: Recommendation, start_row: int) -> int:
        """Build the economic analysis section."""
        row = start_row

        # Section header
        self._merge_and_write(
            row, 1, 6, "ECONOMIC ANALYSIS", self.section_font, alignment=self.left_alignment
        )
        row += 1

        # Economic data table
        headers = ["Metric", "Value", "Explanation"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        row += 1

        # CAPEX
        capex_value = self._format_currency(recommendation.capex_total_inr)
        capex_explanation = f"Upfront investment required for equipment and installation"
        self._write_data_row(row, "Capital Expenditure (CAPEX)", capex_value, capex_explanation)
        row += 1

        # OPEX
        opex_value = self._format_currency(recommendation.annual_opex_inr)
        opex_explanation = f"Annual operating costs including maintenance, fuel, and electricity"
        self._write_data_row(row, "Annual Operating Expenditure (OPEX)", opex_value, opex_explanation)
        row += 1

        # Payback
        payback_value = f"{self._format_years(recommendation.payback_range_years[0])} to {self._format_years(recommendation.payback_range_years[1])}"
        payback_explanation = f"Time required to recover the initial investment through energy cost savings"
        self._write_data_row(row, "Payback Period", payback_value, payback_explanation)
        row += 2

        return row

    def _build_environmental_section(self, recommendation: Recommendation, start_row: int) -> int:
        """Build the environmental impact section."""
        row = start_row

        # Section header
        self._merge_and_write(
            row, 1, 6, "ENVIRONMENTAL IMPACT", self.section_font, alignment=self.left_alignment
        )
        row += 1

        # Environmental data table
        headers = ["Metric", "Value", "Explanation"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        row += 1

        # CO2 reduction
        co2_value = self._format_percentage(recommendation.co2_reduction_pct)
        co2_explanation = f"Reduction in carbon dioxide emissions compared to current operations"
        self._write_data_row(row, "CO₂ Emissions Reduction", co2_value, co2_explanation)
        row += 1

        # Fossil fuel reduction
        fossil_value = self._format_percentage(recommendation.fossil_fuel_reduction_pct)
        fossil_explanation = f"Reduction in fossil fuel consumption, improving energy security"
        self._write_data_row(row, "Fossil Fuel Reduction", fossil_value, fossil_explanation)
        row += 2

        return row

    def _build_policy_section(self, recommendation: Recommendation, start_row: int) -> int:
        """Build the policy and financing section with disclaimer."""
        row = start_row

        # Section header
        self._merge_and_write(
            row, 1, 6, "GOVERNMENT FINANCING & POLICY SUPPORT", self.section_font, alignment=self.left_alignment
        )
        row += 1

        policy = recommendation.explanation.policy_benefits

        if policy.eligible_schemes:
            # Eligible schemes
            schemes_text = f"Eligible Schemes ({len(policy.eligible_schemes)}): {', '.join(policy.eligible_schemes[:5])}"
            if len(policy.eligible_schemes) > 5:
                schemes_text += f" and {len(policy.eligible_schemes) - 5} others"
            self.ws.cell(row=row, column=1, value="Eligible Schemes:")
            self.ws.cell(row=row, column=2, value=schemes_text)
            self.ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            # Total benefit with verification status
            benefit_value = self._format_currency(policy.estimated_total_benefit_inr)
            verification_status = "Verified" if policy.total_benefit_verified else "Unverified"
            self.ws.cell(row=row, column=1, value="Estimated Total Benefit:")
            self.ws.cell(row=row, column=2, value=f"{benefit_value} ({verification_status})")
            self.ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            # Disclaimer - prominently displayed as warning row
            if not policy.total_benefit_verified and policy.disclaimer:
                self._merge_and_write(
                    row, 1, 6,
                    f"IMPORTANT: {policy.disclaimer}",
                    self.disclaimer_font, self.warning_fill, self.left_alignment
                )
                row += 2
        else:
            self.ws.cell(row=row, column=1, value="No eligible government schemes identified for this scenario.")
            self.ws.cell(row=row, column=1).font = Font(italic=True)
            row += 2

        return row

    def _build_sensitivity_section(self, recommendation: Recommendation, start_row: int) -> int:
        """Build the sensitivity analysis section."""
        row = start_row

        # Section header
        self._merge_and_write(
            row, 1, 6, "RISK & SENSITIVITY ANALYSIS", self.section_font, alignment=self.left_alignment
        )
        row += 1

        sensitivity = recommendation.explanation.sensitivity_notes

        # Payback range table
        headers = ["Metric", "Value", "Explanation"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        row += 1

        # P10, P50, P90
        self._write_data_row(
            row, "Optimistic (P10)", self._format_years(sensitivity.payback_p10_years),
            "Best-case payback under favourable conditions"
        )
        row += 1

        self._write_data_row(
            row, "Expected (P50)", self._format_years(sensitivity.payback_p50_years),
            "Most likely payback period"
        )
        row += 1

        self._write_data_row(
            row, "Conservative (P90)", self._format_years(sensitivity.payback_p90_years),
            "Worst-case payback under adverse conditions"
        )
        row += 1

        self._write_data_row(
            row, "Uncertainty Spread", f"{sensitivity.spread_ratio:.2f}",
            "Measure of payback variability (higher = more uncertainty)"
        )
        row += 1

        # Risk interpretation
        self.ws.cell(row=row, column=1, value="Risk Assessment:")
        self.ws.cell(row=row, column=2, value=sensitivity.risk_interpretation)
        self.ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # Top risk factors
        if sensitivity.top_risk_factors:
            factors_text = ", ".join(sensitivity.top_risk_factors)
            self.ws.cell(row=row, column=1, value="Key Risk Factors:")
            self.ws.cell(row=row, column=2, value=factors_text)
            self.ws.cell(row=row, column=1).font = Font(bold=True)
            row += 2

        return row

    def _build_scenario_comparison_table(
        self, recommendation: Recommendation, optimization_result: OptimizationResult, start_row: int
    ) -> int:
        """Build the scenario comparison table with multiple technologies side by side."""
        row = start_row

        # Section header
        self._merge_and_write(
            row, 1, 6, "SCENARIO COMPARISON TABLE", self.section_font, alignment=self.left_alignment
        )
        row += 1

        # Comparison table headers
        headers = ["Rank", "Scenario", "Technologies", "Cost Score", "Emissions Score", "Risk Score", "Overall Score"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        row += 1

        # Add each scenario as a row
        for ranked_scenario in optimization_result.ranked_scenarios:
            tech_sequence = " → ".join(ranked_scenario.technology_sequence)

            self.ws.cell(row=row, column=1, value=ranked_scenario.rank)
            self.ws.cell(row=row, column=2, value=ranked_scenario.scenario_id)
            self.ws.cell(row=row, column=3, value=tech_sequence)
            self.ws.cell(row=row, column=4, value=f"{ranked_scenario.objective_scores.get('cost', 0):.2f}")
            self.ws.cell(row=row, column=5, value=f"{ranked_scenario.objective_scores.get('emissions', 0):.2f}")
            self.ws.cell(row=row, column=6, value=f"{ranked_scenario.objective_scores.get('risk', 0):.2f}")
            self.ws.cell(row=row, column=7, value=f"{ranked_scenario.composite_score:.2f}")

            # Highlight recommended scenario
            if ranked_scenario.scenario_id == recommendation.recommended_scenario_id:
                for col in range(1, 8):
                    self.ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="90EE90", end_color="90EE90", fill_type="solid"
                    )
                    self.ws.cell(row=row, column=col).border = self.thin_border
            else:
                for col in range(1, 8):
                    self.ws.cell(row=row, column=col).border = self.thin_border

            row += 1

        # Add explanation note
        note = (
            "Note: Scores range from 0 to 1 (higher is better). Cost score reflects economic efficiency, "
            "emissions score shows environmental benefit, and risk score indicates operational reliability. "
            "The recommended scenario is highlighted in green."
        )
        self._merge_and_write(row, 1, 7, note, Font(italic=True, size=9), alignment=self.left_alignment)
        row += 2

        return row

    def _build_alternatives_section(self, recommendation: Recommendation, start_row: int) -> int:
        """Build the alternatives analysis section."""
        row = start_row

        # Section header
        self._merge_and_write(
            row, 1, 6, "WHY OTHER OPTIONS WERE NOT RECOMMENDED", self.section_font, alignment=self.left_alignment
        )
        row += 1

        # Alternatives table
        headers = ["Technology Pathway", "Rank", "Score", "Reason", "Key Weakness"]
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        row += 1

        # Add each rejected scenario
        for rejected in recommendation.explanation.why_others_rejected:
            tech_sequence = " → ".join(rejected.technology_sequence)

            self.ws.cell(row=row, column=1, value=tech_sequence)
            self.ws.cell(row=row, column=2, value=rejected.rank)
            self.ws.cell(row=row, column=3, value=f"{rejected.composite_score:.2f}")
            self.ws.cell(row=row, column=4, value=rejected.reason)
            self.ws.cell(row=row, column=5, value=rejected.key_weakness)

            for col in range(1, 6):
                self.ws.cell(row=row, column=col).border = self.thin_border

            row += 1

        row += 1
        return row

    def _build_disclaimer_section(self, recommendation: Recommendation, start_row: int) -> int:
        """Build the final disclaimer section."""
        row = start_row

        # Final disclaimer
        disclaimer_text = (
            "Report Disclaimer: This recommendation is based on the available data and assumptions "
            "documented in the knowledge base. Actual results may vary based on site-specific conditions, "
            "market fluctuations, and policy changes. Consult with qualified engineers and financial "
            "advisors before making investment decisions."
        )

        self._merge_and_write(
            row, 1, 6,
            disclaimer_text,
            self.disclaimer_font, alignment=self.left_alignment
        )

        return row + 1

    def _write_data_row(self, row: int, metric: str, value: str, explanation: str):
        """Write a data row with metric, value, and explanation."""
        self.ws.cell(row=row, column=1, value=metric)
        self.ws.cell(row=row, column=1).font = Font(bold=True)
        self.ws.cell(row=row, column=2, value=value)
        self.ws.cell(row=row, column=3, value=explanation)

        for col in range(1, 4):
            self.ws.cell(row=row, column=col).border = self.thin_border


def generate_excel_report(
    recommendation: Recommendation,
    optimization_result: Optional[OptimizationResult] = None,
    output_path: Optional[Path] = None,
) -> Path:
    """
    Generate an Excel report from a Recommendation object.

    Parameters
    ----------
    recommendation : Recommendation
        The recommendation object to convert to Excel
    optimization_result : OptimizationResult, optional
        Full optimization result for scenario comparison table
    output_path : Path, optional
        Path for the output Excel file. If None, uses a default name.

    Returns
    -------
    Path
        Path to the generated Excel file
    """
    if output_path is None:
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        output_path = Path(
            f"recommendation_{recommendation.factory_id}_{timestamp}.xlsx"
        )

    generator = ExcelReportGenerator(output_path)
    return generator.generate(recommendation, optimization_result)